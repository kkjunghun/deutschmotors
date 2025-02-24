import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook

def upload_excel_files():
    """ Streamlit UI에서 다중 엑셀 파일을 업로드하는 함수 """
    return st.file_uploader("📂 엑셀 파일을 선택하세요", type=["xlsx"], accept_multiple_files=True)

def merge_excel_files(uploaded_files):
    """ 업로드된 다수의 엑셀 파일을 하나의 파일로 병합 """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for file in uploaded_files:
            file_name = file.name.split('.')[0]  # 파일명에서 확장자 제거
            xls = pd.ExcelFile(file, engine='openpyxl')  # openpyxl로 엑셀 파일 로드
            
            for sheet_name in xls.sheet_names:
                sheet_df = pd.read_excel(xls, sheet_name=sheet_name, engine='openpyxl')
                
                # 엑셀 파일의 서식을 복사하기 위한 작업
                wb = load_workbook(file)
                sheet = wb[sheet_name]
                
                new_sheet_name = f"{file_name}"  # 파일명_원래시트명 형식
                sheet_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                
                # 시트의 열 너비, 행 높이, 숫자 표기법, 날짜 표기법 복사
                worksheet = writer.sheets[new_sheet_name]
                
                # 열 너비 자동 조정
                for col in sheet.columns:
                    column = col[0].column_letter  # 열 번호 (A, B, C, ...)
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)  # 여유 공간을 위해 2 추가
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # 행 높이 복사
                for row in sheet.iter_rows():
                    row_height = sheet.row_dimensions[row[0].row].height
                    worksheet.row_dimensions[row[0].row].height = row_height
                
                # 숫자 표기법 및 날짜 표기법 복사
                for row in sheet.iter_rows():
                    for cell in row:
                        new_cell = worksheet[cell.coordinate]
                        if cell.number_format:
                            new_cell.number_format = cell.number_format

    output.seek(0)
    return output

def run_excel_merge():
    """ Streamlit에서 엑셀 병합 기능 실행 """
    st.title("엑셀 파일 병합기")
    st.write("다수의 엑셀 파일을 업로드하여 하나의 파일로 병합합니다. 각 파일의 내용은 파일명과 동일한 시트명으로 저장됩니다.")

    uploaded_files = upload_excel_files()
    
    if not uploaded_files:
        st.warning("⚠️ 하나 이상의 엑셀 파일을 업로드해주세요.")
        return
    
    st.success(f"{len(uploaded_files)}개의 파일이 업로드되었습니다.")  # 업로드된 파일 개수 확인
    
    # 병합된 엑셀 파일 생성
    merged_file = merge_excel_files(uploaded_files)

    # 다운로드 버튼 추가
    st.download_button(
        label="📥 병합된 엑셀 파일 다운로드",
        data=merged_file,
        file_name="merged_excel.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
