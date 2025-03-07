import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import tempfile
import shutil
import time

def apply_excel_date_format(file_path, date_columns):
    """ 엑셀 파일의 날짜 컬럼을 'YYYY-MM-DD' 형식으로 변경하는 함수 """
    wb = load_workbook(file_path)  # 엑셀 파일 로드
    for sheet in wb.sheetnames:  # 모든 시트에 대해 적용
        ws = wb[sheet]
        
        # 날짜 서식 스타일 생성 (YYYY-MM-DD)
        date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
        if "datetime" not in wb.named_styles:
            wb.add_named_style(date_style)

        # 날짜 컬럼 찾아서 스타일 적용
        for col_idx, col in enumerate(ws.iter_cols(), start=1):
            col_letter = get_column_letter(col_idx)
            if ws[col_letter + "1"].value in date_columns:  # 첫 번째 행이 컬럼명
                for cell in col[1:]:  # 첫 번째 행 제외하고 적용
                    if isinstance(cell.value, datetime):
                        cell.style = date_style
    
    wb.save(file_path)  # 적용된 파일 저장

def get_date_info():
    """현재 날짜를 기준으로 전월, 당월, 전월의 마지막 날을 계산하는 함수"""
    today = datetime.today()
    current_month = today.strftime("%Y-%m")
    previous_month = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    previous_month_last_day = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m-%d")
    
    return current_month, previous_month, previous_month_last_day

# 날짜 정보 가져오기
current_month, previous_month, previous_month_last_day = get_date_info()

def get_analysis_settings():
    """ 분석에 필요한 날짜 컬럼과 사원 구분 리스트 반환 """
    date_columns = ["입사일", "퇴사일"]
    employee_types = ["정규직", "계약직", "파견직", "임원"]  # 가나다순 정렬
    
    return date_columns, employee_types

# 분석 대상 컬럼 및 사원 구분 정보 가져오기
date_columns, employee_types = get_analysis_settings()

def get_sheet_order():
    """ 엑셀 병합 시 사용할 시트 정렬 순서를 반환하는 함수 """
    return [
        "도이치아우토",
        "브리티시오토",
        "바이에른오토",
        "이탈리아오토모빌리",
        "브리타니아오토",
        "디티네트웍스",
        "DT네트웍스",
        "도이치파이낸셜",
        "BAMC",
        "차란차",
        "디티이노베이션",
        "DT이노베이션",
        "도이치오토월드",
        "DAFS",
        "사직오토랜드"
    ]
# 시트 정렬 순서 가져오기
sheet_order = get_sheet_order()

def select_month():
    """ Streamlit UI에서 기준 연도 및 월을 선택하는 함수 """
    st.sidebar.subheader("📅 기준 월 설정")
    selected_year = st.sidebar.selectbox("📌 기준 연도 선택", list(range(2022, datetime.today().year + 1)), index=2)
    selected_month = st.sidebar.selectbox("📌 기준 월 선택", list(range(1, 13)), index=datetime.today().month - 2)

    # 선택한 기준 월을 YYYY-MM 형식으로 변환
    selected_date = datetime(selected_year, selected_month, 1)
    selected_month_str = selected_date.strftime("%Y-%m")  # 기준 월 (예: 2023-11)
    selected_month_last_day = (selected_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)  # 기준 월의 마지막 날

    st.sidebar.write(f"📌 선택된 기준 월: **{selected_month_str}**")

    return selected_month_str, selected_month_last_day


def select_feature():
    """ Streamlit UI에서 사용자가 사용할 기능을 선택하는 함수 """
    return st.sidebar.selectbox(
        "📌 사용할 기능을 선택하세요",
        ["엑셀 병합 및 인원 분석", "4대보험료 검증 시스템", "급여 업무 시스템", "채용 분석 시스템"]
    )

def get_delete_keywords():
    """ Streamlit UI에서 키워드 기반 삭제 컬럼을 입력받는 함수 """
    st.sidebar.subheader("🔒 개인정보 보호 설정")

    # 사용자가 쉼표로 구분하여 키워드를 입력하면 리스트로 변환
    delete_keywords_input = st.sidebar.text_area("🔍 키워드로 삭제할 컬럼 입력 (쉼표로 구분)", "주민, 경력, 인정")
    delete_keywords = [kw.strip() for kw in delete_keywords_input.split(",") if kw.strip()]
    
    return delete_keywords


def upload_excel_files():
    """ Streamlit UI에서 다중 엑셀 파일을 업로드하는 함수 """
    return st.file_uploader("📂 엑셀 파일을 선택하세요", type=["xlsx"], accept_multiple_files=True)



def save_uploaded_files(uploaded_files):
    """ 업로드된 엑셀 파일을 임시 폴더에 저장하는 함수 """
    temp_dir = tempfile.mkdtemp()  # 임시 폴더 생성
    merged_excel_path = os.path.join(temp_dir, "merged_excel.xlsx")  # 병합된 파일 저장 경로

    file_paths = []
    for uploaded_file in uploaded_files:
        file_path = os.path.join(temp_dir, uploaded_file.name)
        with open(file_path, "wb") as f:
            f.write(uploaded_file.read())
        file_paths.append(file_path)
    
    return temp_dir, merged_excel_path, file_paths

# 📌 엑셀 병합 함수 실행
def merge_excel_files(files, output_file, sheet_order, delete_keywords):
    """ 여러 개의 엑셀 파일을 병합하고, 특정 키워드가 포함된 컬럼을 삭제하는 함수 """
    
    # 시트 정렬 순서에 따라 정렬
    files.sort(key=lambda x: sheet_order.index(os.path.splitext(os.path.basename(x))[0]) if os.path.splitext(os.path.basename(x))[0] in sheet_order else len(sheet_order))

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for file in files:
            try:
                wb = load_workbook(file, data_only=True)
                sheet_names = wb.sheetnames  

                if not sheet_names:
                    st.warning(f"⚠️ 파일 `{os.path.basename(file)}` 에 사용 가능한 시트가 없어 건너뜁니다.")
                    continue

                for sheet_name in sheet_names:
                    ws = wb[sheet_name]
                    data = [[cell.value for cell in row] for row in ws.iter_rows()]
                    
                    if not data or all(all(cell is None for cell in row) for row in data):
                        st.warning(f"⚠️ 파일 `{os.path.basename(file)}` 의 시트 `{sheet_name}` 가 비어 있어 건너뜁니다.")
                        continue

                    header_row_index = None
                    for idx, row in enumerate(data):
                        if row[0] == "No":
                            header_row_index = idx
                            break

                    if header_row_index is not None:
                        headers = data[header_row_index]
                        df = pd.DataFrame(data[header_row_index + 1:], columns=headers)
                    else:
                        df = pd.DataFrame(data[1:], columns=data[0])

                    # 컬럼명 공백 제거
                    df.columns = df.columns.str.strip()

                    # ✅ **키워드 기반 삭제 처리**
                    delete_cols_by_keyword = [col for col in df.columns if any(keyword in col for keyword in delete_keywords)]
                    
                    # 컬럼 삭제 (키워드 포함 컬럼만 삭제)
                    df.drop(columns=[col for col in delete_cols_by_keyword if col in df.columns], errors="ignore", inplace=True)

                    # 시트 이름이 31자를 초과하지 않도록 잘라서 저장
                    sheet_name_trimmed = os.path.splitext(os.path.basename(file))[0][:31]
                    df.to_excel(writer, sheet_name=sheet_name_trimmed, index=False)

            except Exception as e:
                st.error(f"🚨 파일 `{os.path.basename(file)}` 처리 중 오류 발생: {e}")


def process_employee_data(df, sheet_name, selected_month_str, previous_month, previous_month_last_day, date_columns):
    """
    직원 데이터를 정리하고 입사자, 퇴사자, 재직자 수 등을 계산하는 함수
    """
    # 📌 컬럼명 정리
    if "Starting Date" in df.columns:
        df.rename(columns={"Starting Date": "입사일"}, inplace=True)
    df.columns = df.columns.str.strip()

    # 📌 특정 인원 제외
    exclude_conditions = {
        "도이치오토월드": ["장준호"],
        "DT네트웍스": ["권혁민"],
        "디티네트웍스": ["권혁민"],
        "BAMC": ["YOON JONG LYOL"]
    }
    if sheet_name in exclude_conditions and "성명" in df.columns:
        df = df.loc[~df["성명"].isin(exclude_conditions[sheet_name])]
    if "English Name" in df.columns:
        df = df.loc[~df["English Name"].isin(exclude_conditions[sheet_name])]

    # 📌 날짜 변환
    if "입사일" in df.columns:
        df["입사일"] = pd.to_datetime(df["입사일"], errors="coerce").dt.strftime("%Y-%m-%d")
    if "퇴사일" not in df.columns:
        df["퇴사일"] = None
    if "Remark" in df.columns:
        df.loc[df["Remark"].astype(str).str.startswith("Resigned and last working"), "퇴사일"] = previous_month_last_day

    # 📌 "사원구분명" 컬럼 자동 생성
    if "사원구분명" not in df.columns:
        df["사원구분명"] = None
    if "Contract Type" in df.columns:
        df.loc[df["Contract Type"].astype(str).str.contains("FDC", na=False), "사원구분명"] = "계약직"
        df.loc[df["Contract Type"].astype(str).str.contains("UDC", na=False), "사원구분명"] = "정규직"

    # 📌 날짜 변환 (YYYY-MM)
    for col in date_columns:
        df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m")

    # 📌 원하는 정렬 순서 지정
    employee_type_order = ["임원", "정규직", "계약직", "파견직"]

    # 📌 1. 선택한 월 입사자 수
    new_hires_selected_month = df[df["입사일"] == selected_month_str].shape[0]
    st.write(f"📌 1. **{selected_month_str} 입사자 수:** {new_hires_selected_month}명")

    # 📌 2. 선택한 월 퇴사자 수
    resigned_selected_month = df[df["퇴사일"] == selected_month_str].shape[0]
    st.write(f"📌 2. **{selected_month_str} 퇴사자 수:** {resigned_selected_month}명")

    # 📌 3. 선택한 월 기준 총 재직자 수
    active_this_month = df[
        (df["입사일"] <= selected_month_str) & 
        (df["퇴사일"].isna() | (df["퇴사일"] > selected_month_str))
    ].shape[0]
    st.write(f"📌 3. **{selected_month_str} 기준 총 재직자 수:** {active_this_month}명")

    # 📌 4. 선택한 월 입사자 수 (사원구분별)
    new_hires_by_type_selected_month = df[df["입사일"] == selected_month_str]["사원구분명"].value_counts()
    new_hires_by_type_selected_month = new_hires_by_type_selected_month.reindex(employee_type_order, fill_value=0)
    st.write(f"📌 4. **{selected_month_str} 입사자 수 (사원구분별)**")
    for emp_type, count in new_hires_by_type_selected_month.items():
        st.write(f"  - {emp_type}: {count}명")

    # 📌 5. 선택한 월 퇴사자 수 (사원구분별)
    resigned_by_type_selected_month = df[df["퇴사일"] == selected_month_str]["사원구분명"].value_counts()
    resigned_by_type_selected_month = resigned_by_type_selected_month.reindex(employee_type_order, fill_value=0)
    st.write(f"📌 5. **{selected_month_str} 퇴사자 수 (사원구분별)**")
    for emp_type, count in resigned_by_type_selected_month.items():
        st.write(f"  - {emp_type}: {count}명")

    # 📌 6. 선택한 월 기준 재직자 수 (사원구분별)
    active_this_month_by_type = df[
        (df["입사일"] <= selected_month_str) & 
        (df["퇴사일"].isna() | (df["퇴사일"] > selected_month_str))
    ]["사원구분명"].value_counts()
    active_this_month_by_type = active_this_month_by_type.reindex(employee_type_order, fill_value=0)
    st.write(f"📌 6. **{selected_month_str} 기준 총 재직자 수 (사원구분별)**")
    for emp_type, count in active_this_month_by_type.items():
        st.write(f"  - {emp_type}: {count}명")

    # 📌 입사자 및 퇴사자 정보 저장
    all_new_hires = []
    all_resigned = []

    if {"입사일", "사원구분명", "부서명", "성명", "직급명"}.issubset(df.columns):
        new_hires = df[df["입사일"] == previous_month][["사원구분명", "부서명", "성명", "직급명"]]
        if not new_hires.empty:
            new_hires["시트명"] = sheet_name
            all_new_hires.append(new_hires)

    if {"퇴사일", "사원구분명", "부서명", "성명", "직급명"}.issubset(df.columns):
        resigned = df[df["퇴사일"] == previous_month][["사원구분명", "부서명", "성명", "직급명"]]
        if not resigned.empty:
            resigned["시트명"] = sheet_name
            all_resigned.append(resigned)

    return all_new_hires, all_resigned



def analyze_employee_data(merged_excel_path, selected_month_str, previous_month, previous_month_last_day, date_columns):
    with pd.ExcelWriter(merged_excel_path, engine="openpyxl", mode="a") as writer:
        sheets = pd.read_excel(merged_excel_path, sheet_name=None, engine="openpyxl")

        all_new_hires = []
        all_resigned = []

        for sheet_name, df in sheets.items():
            st.subheader(f"📄 시트 이름: {sheet_name}")

            new_hires, resigned = process_employee_data(df, sheet_name, selected_month_str, previous_month, previous_month_last_day, date_columns)

            if new_hires:
                all_new_hires.extend(new_hires)
            if resigned:
                all_resigned.extend(resigned)

        # 📌 입사자 및 퇴사자 데이터를 엑셀 시트에 저장
        if all_new_hires:
            pd.concat(all_new_hires).to_excel(writer, sheet_name="입사자_리스트", index=False)
        if all_resigned:
            pd.concat(all_resigned).to_excel(writer, sheet_name="퇴사자_리스트", index=False)



def download_excel_file(file_path, temp_dir, file_name="merged_excel.xlsx"):
    """ 병합된 엑셀 파일을 다운로드할 수 있도록 제공하는 함수 """
    if st.download_button(
        label="📥 병합된 엑셀 다운로드",
        data=open(file_path, "rb").read(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        # 일정 시간 후 파일 및 임시 폴더 삭제
        time.sleep(10)
        os.remove(file_path)
        shutil.rmtree(temp_dir)  
        st.warning("🔒 다운로드 후 10초가 지나 파일이 자동 삭제되었습니다.")

def apply_date_format_to_excel(file_path, date_columns):
    """ 병합된 엑셀 파일의 날짜 컬럼을 YYYY-MM-DD 형식으로 변환하는 함수 """
    apply_excel_date_format(file_path, date_columns)

def process_excel_files(uploaded_files, selected_month_str, previous_month, previous_month_last_day, date_columns, sheet_order, delete_keywords):
    """ 엑셀 파일을 병합, 분석, 서식 적용 후 다운로드할 수 있도록 처리하는 함수 """
    
    # 📌 1. 업로드된 파일을 저장
    temp_dir, merged_excel_path, file_paths = save_uploaded_files(uploaded_files)
    
    # 📌 2. 엑셀 병합 및 키워드 기반 컬럼 삭제
    merge_excel_files(file_paths, merged_excel_path, sheet_order, delete_keywords)
    
    # 📌 3. 병합된 데이터에서 입사자 및 퇴사자 분석
    analyze_employee_data(merged_excel_path, selected_month_str, previous_month, previous_month_last_day, date_columns)
    
    # 📌 4. 날짜 형식 적용
    apply_date_format_to_excel(merged_excel_path, date_columns)
    
    # 📌 5. 다운로드 버튼 제공
    download_excel_file(merged_excel_path, temp_dir)


def run_excel_analysis():
    """ Streamlit UI에서 사용자의 입력을 받고 엑셀 병합 및 분석을 실행하는 함수 """

    # 📌 타이틀 설정
    st.title("📊 다중 엑셀 분석 시스템")

    # ✅ 기능 선택
    feature_option = select_feature()

    # ✅ 기준 월 선택
    selected_month_str, selected_month_last_day = select_month()

    if feature_option == "엑셀 병합 및 인원 분석":
        st.write("엑셀 파일을 업로드하면 자동으로 병합 후 분석을 수행합니다.")

        # ✅ 개인정보 보호 설정 (삭제할 키워드 입력)# 삭제할 키워드 리스트 가져오기
        delete_keywords = get_delete_keywords()

        # ✅ 다중 엑셀 파일 업로드 # 엑셀 파일 업로드 함수 호출
        uploaded_files = upload_excel_files()

        if uploaded_files:
            # ✅ # 전체 엑셀 처리 함수 호출 (한 번에 실행)
            process_excel_files(uploaded_files, selected_month_str, previous_month, previous_month_last_day, date_columns, sheet_order, delete_keywords)

if __name__ == "__main__":
    # Streamlit UI 실행 함수 호출
    run_excel_analysis()
