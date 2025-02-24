import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import tempfile
import shutil
import time

def upload_insurance_files():
    """ Streamlit UI에서 4대보험 데이터 엑셀 파일을 업로드하는 함수 """
    return st.file_uploader(
        "📂 4대보험 데이터 엑셀 파일을 업로드하세요 (다중 선택 가능)", 
        type=["xlsx"], 
        accept_multiple_files=True
    )

def save_uploaded_insurance_files(uploaded_files):
    """ 업로드된 4대보험 엑셀 파일을 임시 폴더에 저장하는 함수 """
    temp_dir = tempfile.mkdtemp()
    merged_excel_path = os.path.join(temp_dir, "merged_insurance_data.xlsx")  # 병합 파일 경로
    
    file_paths = []
    for uploaded_file in uploaded_files:
        file_path = os.path.join(temp_dir, uploaded_file.name)
        with open(file_path, "wb") as f:
            f.write(uploaded_file.read())
        file_paths.append(file_path)

    return temp_dir, merged_excel_path, file_paths 

def merge_insurance_files(file_paths):
    """ 여러 개의 4대보험 엑셀 파일을 병합하고 서식을 유지하는 함수 """
    
    # 📌 병합을 위한 새로운 워크북 생성
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)  # 기본 시트 제거

    if not file_paths:  # 📌 업로드된 파일이 없는 경우 처리
        st.error("❌ 업로드된 4대보험 데이터 파일이 없습니다.")
        return None

    for file_path in file_paths:
        try:
            source_wb = load_workbook(file_path, data_only=False)  # 수식 유지

            for sheet_name in source_wb.sheetnames:
                source_ws = source_wb[sheet_name]

                # ✅ 이미 존재하는 시트가 있으면 덮어쓰기
                if sheet_name in merged_wb.sheetnames:
                    merged_wb.remove(merged_wb[sheet_name])

                new_ws = merged_wb.create_sheet(title=sheet_name)

                # ✅ 열 너비 유지
                for col in source_ws.column_dimensions:
                    new_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

                # ✅ 행 높이 유지
                for row in source_ws.row_dimensions:
                    new_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

                # ✅ 원본 시트 데이터를 복사 (수식 + 서식 유지 + 검정색 텍스트 적용)
                for row in source_ws.iter_rows():
                    for cell in row:
                        new_cell = new_ws[cell.coordinate]
                        new_cell.value = cell.value  # 수식 또는 값 복사

                        # ✅ 스타일 복사 (서식 유지) + 텍스트 검정색 적용
                        if cell.font:
                            new_cell.font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                underline=cell.font.underline,
                                strike=cell.font.strike,
                                color="000000"  # 모든 텍스트를 검정색으로 설정
                            )
                        if isinstance(cell.fill, PatternFill):  # ✅ 배경색 유지
                            new_cell.fill = PatternFill(
                                fill_type=cell.fill.fill_type,
                                fgColor=cell.fill.fgColor,
                                bgColor=cell.fill.bgColor
                            )
                        if cell.border:
                            new_cell.border = Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            )
                        if cell.alignment:
                            new_cell.alignment = Alignment(
                                horizontal=cell.alignment.horizontal,
                                vertical=cell.alignment.vertical,
                                wrap_text=cell.alignment.wrap_text
                            )

                        # ✅ 1000단위 쉼표 적용 (숫자인 경우만)
                        if isinstance(cell.value, (int, float)) and cell.data_type != "f":  # 수식이 아닌 숫자만 적용
                            new_cell.number_format = "#,##0"  # 1000 단위 콤마 적용

                # ✅ 원본 병합된 셀 유지
                for merged_cell in source_ws.merged_cells.ranges:
                    new_ws.merge_cells(str(merged_cell))

        except Exception as e:
            st.error(f"❌ 파일 `{os.path.basename(file_path)}` 처리 중 오류 발생: {e}")
        
    return merged_wb  # 📌 `Workbook` 객체 반환
        
    
# ✅ 다운로드 버튼 생성
def download_merged_insurance_file(merged_wb, merged_excel_path, temp_dir):
    """ 병합된 4대보험 데이터를 다운로드할 수 있도록 제공하는 함수 """
    if merged_wb is None:
        return  # 병합된 파일이 없으면 실행 중지

    merged_wb.save(merged_excel_path)  # 📌 병합된 엑셀 저장

    with open(merged_excel_path, "rb") as file:
        st.download_button(
            label="📥 병합된 4대보험 데이터 다운로드",
            data=file,
            file_name="merged_insurance_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # ✅ 일정 시간 후 자동 삭제
    time.sleep(10)
    shutil.rmtree(temp_dir)  
    st.warning("🔒 다운로드 후 10초가 지나 파일이 자동 삭제되었습니다.")

# ✅ 4대보험 검증 시스템 실행
def run_insurance_analysis():
    """ 4대보험 검증 시스템 실행 함수 """
    st.subheader("4대보험료 검증 시스템")

    uploaded_insurance_files = upload_insurance_files()

    if uploaded_insurance_files:
        temp_dir, merged_excel_path, file_paths = save_uploaded_insurance_files(uploaded_insurance_files)

        merged_wb = merge_insurance_files(file_paths)

        download_merged_insurance_file(merged_wb, merged_excel_path, temp_dir)



# if feature_option == "4대보험료 검증 시스템":
#     run_insurance_analysis()
